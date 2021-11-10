import smtplib, ssl

import openpyxl
from datetime import date

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


path="Book1.xlsx"
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 
row = sheet_obj.max_row


d=str(date.today())+" 00:00:00"
cell_obj = sheet_obj.cell(row = 2, column = 2)

if(str(d)==str(cell_obj.value)):
    sender_email = "Enter sender's email"
    receiver_email = "Enter receiver's email"
    password = input("Type your password and press enter:")
    
    message = MIMEMultipart("alternative")
    message["Subject"] = "Hey there"
    message["From"] = sender_email
    message["To"] = receiver_email
    
    
    html = """\
    <html>
      <body>
        <p>Hi,<br>
           How are you?
           This is Apoorva
           Today is a holiday!!<br>
           
        </p>
      </body>
    </html>
    """
    

    part1 = MIMEText(text, "plain")
    part2 = MIMEText(html, "html")
    
    
    message.attach(part1)
    message.attach(part2)
    
    # Create secure connection with server and send email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(
            sender_email, receiver_email, message.as_string()
        )
#Less secure apps should be switched on ,in the sender's mail settings before executing the above prgram
